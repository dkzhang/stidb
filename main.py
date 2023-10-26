import json
import os
import shutil
from jinja2 import Environment, DictLoader, Template
from openpyxl import load_workbook
from flask import Flask, send_file, request, jsonify
from openpyxl import load_workbook
from io import BytesIO
from werkzeug.utils import secure_filename

from model.person import Person

app = Flask(__name__)

app.config['UPLOAD_FOLDER'] = os.getcwd()


@app.route('/upload_template', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return {"error": "No file part in the request."}, 400
    file = request.files['file']
    if file.filename == '':
        return {"error": "No selected file."}, 400
    if file:
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        return {"success": "File uploaded successfully."}, 200


@app.route('/to_excel', methods=['POST'])
def modify_excel():
    # with open('person_all_info2.json', 'r', encoding='utf-8') as f:
    #     json_string = f.read()

    try:
        # Get JSON data from the request
        json_string = request.get_data(as_text=True)

        # Parse the JSON string into a Python list
        person_info = Person.from_json(json_string)
        print(person_info)

    except Exception as e:
        return jsonify({"error": str(e)}), 400  # Bad request

    filename = person_info.excel_template
    base, ext = os.path.splitext(filename)
    new_filename = f"{base}_{person_info.department}_{person_info.name}({person_info.sap}){ext}"

    # shutil.copyfile(filename, new_filename)

    wb = load_workbook(filename)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Iterate over the cells
        for row in sheet.iter_rows():
            for cell in row:
                cell_value = str(cell.value)
                print("cell_value = " + cell_value)
                if "{{" in cell_value and "}}" in cell_value:
                    # It is a template, render it
                    t_name = f"template_{cell.row}_{cell.column}"
                    env = Environment(loader=DictLoader({t_name: cell_value}))
                    template = env.get_template(t_name)
                    output = template.render(data=person_info)
                    print(t_name + "output = " + output)
                    cell.value = output

    # Save modified workbook to a BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)

    # Seek to the beginning of the stream
    excel_file.seek(0)

    # Return modified Excel file
    return send_file(excel_file, download_name=new_filename, as_attachment=True)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)

#
# filename = 'data2.xlsx'
# new_filename = 'new_data.xlsx'
#
# shutil.copyfile(filename, new_filename)
#
# wb = load_workbook(new_filename)
#
# for sheet_name in wb.sheetnames:
#     sheet = wb[sheet_name]
#     # Iterate over the cells
#     for row in sheet.iter_rows():
#         for cell in row:
#             cell_value = str(cell.value)
#             print("cell_value = " + cell_value)
#             if "{%" in cell_value and "%}" in cell_value:
#                 # It is a template, render it
#                 t_name = f"template_{cell.row}_{cell.column}"
#                 env = Environment(loader=DictLoader({t_name: cell_value}))
#                 template = env.get_template(t_name)
#                 output = template.render(data=data)
#                 print(t_name + "output = " + output)
#                 cell.value = output
#
# wb.save(new_filename)

# from flask import Flask, send_file
# from openpyxl import load_workbook
# from io import BytesIO
#
# app = Flask(__name__)
#
# @app.route('/modify_excel', methods=['GET'])
# def modify_excel():
#     # Load Excel workbook using openpyxl
#     wb = load_workbook(filename='original.xlsx')
#
#     # Select a worksheet by name
#     ws = wb.get_sheet_by_name('Sheet1')
#
#     # Modify a cell value (this example sets cell A1 value to 'Modified')
#     ws['A1'] = 'Modified'
#
#     # Save modified workbook to a BytesIO object
#     excel_file = BytesIO()
#     wb.save(excel_file)
#
#     # Seek to the beginning of the stream
#     excel_file.seek(0)
#
#     # Return modified Excel file
#     return send_file(excel_file, attachment_filename='modified.xlsx', as_attachment=True)
#
# if __name__ == '__main__':
#     app.run(host='0.0.0.0', port=5000)

# # Define your template
# template_source = """
# {% for item in data if item.id%2 == 0 %}
#     {% if loop.index <= 5 %}
#         Index: {{ loop.index }} - Id: {{ item.id }}, Name: {{ item.name }}
#     {% endif %}
# {% endfor %}
# 123456
# {% for item in data -%}
# Index: {{ loop.index }} - Id: {{ item.id }}
# {% endfor %}
# """
